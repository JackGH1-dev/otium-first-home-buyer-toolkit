#!/usr/bin/env python3
"""
QED Excel Cell Scanner - Find where MAX Loan result is stored
"""

import openpyxl
from pathlib import Path

def scan_for_max_loan():
    """Scan Excel sheet for MAX Loan related cells"""
    
    excel_path = Path(r"C:\Users\encou\Documents\Project MicroSass\Otium\qed_serviceability_calculator_3_28_may_2025_download_450.xlsm")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path, keep_vba=True, data_only=True)
    
    # Test both worksheets
    for worksheet_name in ["Single income", "Dual income"]:
        ws = wb[worksheet_name]
        print(f"\n{'='*60}")
        print(f"Worksheet: {worksheet_name}")
        print(f"{'='*60}")
        
        # Set a simple test scenario
        ws['F8'] = 100000  # Primary income
        if worksheet_name == "Dual income":
            ws['I8'] = 80000  # Secondary income
        ws['E3'].value = 0  # No dependents
        ws['B7'] = 0.055  # Interest rate
        
        # Save and reload to trigger calculations
        temp_path = excel_path.with_suffix('.temp.xlsm')
        wb.save(temp_path)
        wb.close()
        
        # Reload with data_only to get calculated values
        wb = openpyxl.load_workbook(temp_path, keep_vba=True, data_only=True)
        ws = wb[worksheet_name]
        
        print("\nSearching for 'MAX' or 'Loan' labels and nearby values:")
        print("-" * 40)
        
        # Scan for cells containing MAX or Loan text
        max_loan_cells = []
        for row in range(1, 50):
            for col in range(1, 10):  # A to I columns
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value
                if cell_value and isinstance(cell_value, str):
                    if 'MAX' in cell_value.upper() or 'LOAN' in cell_value.upper():
                        cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        # Check adjacent cells for values
                        right_cell = ws.cell(row=row, column=col+1)
                        below_cell = ws.cell(row=row+1, column=col)
                        
                        print(f"\n{cell_ref}: '{cell_value}'")
                        if right_cell.value:
                            print(f"  -> Right cell ({openpyxl.utils.get_column_letter(col+1)}{row}): {right_cell.value}")
                        if below_cell.value:
                            print(f"  v Below cell ({openpyxl.utils.get_column_letter(col)}{row+1}): {below_cell.value}")
        
        print("\n\nScanning F column (rows 40-45) specifically:")
        print("-" * 40)
        for row in range(40, 46):
            cell_ref = f"F{row}"
            cell = ws[cell_ref]
            if cell.value is not None:
                print(f"{cell_ref}: {cell.value} (type: {type(cell.value).__name__})")
        
        print("\n\nScanning E column (rows 40-45) specifically:")
        print("-" * 40)
        for row in range(40, 46):
            cell_ref = f"E{row}"
            cell = ws[cell_ref]
            if cell.value is not None:
                print(f"{cell_ref}: {cell.value} (type: {type(cell.value).__name__})")
        
        # Also scan for large numeric values (likely loan amounts)
        print("\n\nLarge numeric values found (>100,000):")
        print("-" * 40)
        for row in range(1, 50):
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and cell.value > 100000:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    print(f"{cell_ref}: ${cell.value:,.0f}")
    
    wb.close()
    print("\nScan complete!")

if __name__ == "__main__":
    scan_for_max_loan()