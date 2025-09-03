#!/usr/bin/env python3
"""
QED Serviceability Calculator Automated Tester
Tests all 6 borrowing power scenarios automatically
"""

import openpyxl
from pathlib import Path
import json

def test_qed_scenario(scenario, worksheet_name="Dual income"):
    """Test a single scenario in QED calculator"""
    
    excel_path = Path(r"C:\Users\encou\Documents\Project MicroSass\Otium\qed_serviceability_calculator_3_28_may_2025_download_450.xlsm")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, keep_vba=True, data_only=False)
        
        # Select appropriate worksheet
        if scenario['secondary_income'] > 0:
            ws = wb["Dual income"]
        else:
            ws = wb["Single income"]
        
        print(f"Testing: {scenario['name']}")
        print(f"Using worksheet: {ws.title}")
        
        # Clear existing values first
        ws['F8'] = 0  # Primary income
        ws['I8'] = 0  # Secondary income  
        ws['E3'].value = 0  # Dependents
        ws['B7'] = 0.055  # Default rate
        ws['F9'] = "N"  # HECS debt flag primary
        ws['I9'] = "N"  # HECS debt flag secondary
        ws['F16'] = 0  # HECS debt amount primary
        ws['I16'] = 0  # HECS debt amount secondary
        ws['F10'] = 0  # Rental income
        ws['F33'] = 0  # Current rent
        
        # Set input values
        ws['F8'] = scenario['primary_income']  # Primary income
        if scenario['secondary_income'] > 0:
            ws['I8'] = scenario['secondary_income']  # Secondary income
            
        ws['E3'].value = scenario['dependents']  # Dependents
        ws['B7'] = scenario['interest_rate'] / 100  # Interest rate as decimal
        
        # HECS debt setup
        if scenario['hecs_primary'] > 0:
            ws['F9'] = "Y"
            ws['F16'] = scenario['hecs_primary']
        if scenario['hecs_secondary'] > 0:
            ws['I9'] = "Y" 
            ws['I16'] = scenario['hecs_secondary']
            
        # Rental income (annual)
        if scenario['rental_income'] > 0:
            ws['F10'] = scenario['rental_income']
            
        # Current rent (monthly)
        if scenario['current_rent'] > 0:
            ws['F33'] = scenario['current_rent']
            
        # Set a test loan amount in F5 to trigger calculation
        ws['F5'] = 500000  # Set a default loan amount to trigger calc
        
        # Save to trigger recalculation
        temp_path = excel_path.with_suffix('.temp.xlsm')
        wb.save(temp_path)
        wb.close()
        
        # Reload to get calculated values
        wb = openpyxl.load_workbook(temp_path, keep_vba=True, data_only=True)
        ws = wb["Dual income"] if scenario['secondary_income'] > 0 else wb["Single income"]
        
        # Read result - MAX Loan is in different cells for each worksheet
        result = None
        if ws.title == "Single income":
            result_cells = ['F42']  # Single income: MAX Loan is in F42
        else:
            result_cells = ['F43']  # Dual income: MAX Loan is in F43
        
        for cell_ref in result_cells:
            cell_value = ws[cell_ref].value
            
            # Handle different value types
            if isinstance(cell_value, (int, float)) and cell_value > 0:
                result = cell_value
                print(f"  Found numeric result in {cell_ref}: ${result:,.0f}")
                break
            elif isinstance(cell_value, str) and cell_value.replace('$', '').replace(',', '').replace(' ', '').isdigit():
                # Handle formatted currency strings
                result = float(cell_value.replace('$', '').replace(',', '').replace(' ', ''))
                print(f"  Found formatted result in {cell_ref}: ${result:,.0f}")
                break
            elif cell_value and str(cell_value) not in ['Max loan', 'Loan amount', 'None', '0']:
                print(f"  Checking {cell_ref}: {cell_value} (type: {type(cell_value)})")
                
        if result is None or result == 0:
            print(f"  WARNING: No valid result found. Checking all cells...")
            # Last resort - scan for large numbers
            for row in range(1, 50):
                for col in range(1, 20):
                    cell_value = ws.cell(row=row, column=col).value
                    if isinstance(cell_value, (int, float)) and 100000 < cell_value < 10000000:
                        result = cell_value
                        cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        print(f"  Found large number in {cell_ref}: ${result:,.0f}")
                        break
                if result:
                    break
        
        # Close the workbook
        wb.close()
        
        return {
            'scenario_name': scenario['name'],
            'qed_result': result,
            'our_app_result': scenario['our_app_result'],
            'expected_min': scenario['expected_range'][0],
            'expected_max': scenario['expected_range'][1],
            'worksheet_used': ws.title
        }
        
    except Exception as e:
        print(f"ERROR testing {scenario['name']}: {str(e)}")
        return None

def run_all_scenarios():
    """Run all 6 test scenarios"""
    
    scenarios = [
        {
            'name': 'Scenario 1: Single, Low Income, Owner-Occupied',
            'primary_income': 65000,
            'secondary_income': 0,
            'dependents': 0,
            'hecs_primary': 0,
            'hecs_secondary': 0,
            'property_type': 'Owner-Occupied',
            'interest_rate': 5.5,
            'location': 'NSW 2000',
            'rental_income': 0,
            'current_rent': 650,
            'expected_range': (350000, 400000),
            'our_app_result': 312530
        },
        {
            'name': 'Scenario 2: Single, Medium Income, Investment',
            'primary_income': 95000,
            'secondary_income': 0,
            'dependents': 0,
            'hecs_primary': 25000,
            'hecs_secondary': 0,
            'property_type': 'Investment',
            'interest_rate': 5.8,
            'location': 'VIC 3000',
            'rental_income': 450 * 52,  # Weekly to annual
            'current_rent': 0,
            'expected_range': (600000, 700000),
            'our_app_result': 502553
        },
        {
            'name': 'Scenario 3: Couple, High Income, Owner-Occupied',
            'primary_income': 120000,
            'secondary_income': 85000,
            'dependents': 2,
            'hecs_primary': 35000,
            'hecs_secondary': 20000,
            'property_type': 'Owner-Occupied',
            'interest_rate': 5.5,
            'location': 'QLD 4000',
            'rental_income': 0,
            'current_rent': 650,
            'expected_range': (900000, 1000000),
            'our_app_result': 237413
        },
        {
            'name': 'Scenario 4: Couple, High Income, Investment',
            'primary_income': 140000,
            'secondary_income': 75000,
            'dependents': 0,
            'hecs_primary': 0,
            'hecs_secondary': 0,
            'property_type': 'Investment',
            'interest_rate': 5.8,
            'location': 'WA 6000',
            'rental_income': 650 * 52,  # Weekly to annual
            'current_rent': 400 * 52 / 12,  # Weekly to monthly
            'expected_range': (1200000, 1500000),
            'our_app_result': 1060237
        },
        {
            'name': 'Scenario 5: Single, Very High Income, Investment',
            'primary_income': 180000,
            'secondary_income': 0,
            'dependents': 1,
            'hecs_primary': 45000,
            'hecs_secondary': 0,
            'property_type': 'Investment',
            'interest_rate': 5.8,
            'location': 'SA 5000',
            'rental_income': 800 * 52,  # Weekly to annual
            'current_rent': 0,
            'expected_range': (1500000, 2000000),
            'our_app_result': 660016
        },
        {
            'name': 'Scenario 6: Young Couple, Entry Level',
            'primary_income': 75000,
            'secondary_income': 60000,
            'dependents': 0,
            'hecs_primary': 15000,
            'hecs_secondary': 18000,
            'property_type': 'Owner-Occupied',
            'interest_rate': 5.5,
            'location': 'NSW 2650',
            'rental_income': 0,
            'current_rent': 650,
            'expected_range': (650000, 750000),
            'our_app_result': 425720
        }
    ]
    
    print("QED Automated Testing - All 6 Scenarios")
    print("=" * 60)
    
    results = []
    for scenario in scenarios:
        result = test_qed_scenario(scenario)
        if result:
            results.append(result)
            try:
                qed_result = float(result['qed_result']) if result['qed_result'] else 0
                print(f"  QED Result: ${qed_result:,.0f}")
                print(f"  Our App:    ${result['our_app_result']:,.0f}")
                
                # Calculate variance
                if qed_result > 0:
                    variance = ((result['our_app_result'] - qed_result) / qed_result) * 100
                    print(f"  Variance:   {variance:+.1f}%")
            except (ValueError, TypeError) as e:
                print(f"  QED Result: {result['qed_result']}")
                print(f"  Our App:    ${result['our_app_result']:,.0f}")
                print(f"  Error calculating variance: {e}")
            print()
        else:
            print(f"  FAILED to test scenario")
            print()
    
    # Save results
    results_file = Path(r"C:\Users\encou\Documents\Project MicroSass\Otium\docs\qed_test_results.json")
    with open(results_file, 'w') as f:
        json.dump(results, f, indent=2)
    
    print(f"Results saved to: {results_file}")
    
    # Generate summary
    print("\nSUMMARY:")
    print("-" * 40)
    for result in results:
        if result['qed_result']:
            qed_result = float(result['qed_result']) if result['qed_result'] else 0
            if qed_result > 0:
                variance = ((result['our_app_result'] - qed_result) / qed_result) * 100
                status = "MATCH" if abs(variance) <= 5 else "VARIANCE"
                print(f"{result['scenario_name'][:30]:30} | {variance:+6.1f}% | {status}")
    
    return results

if __name__ == "__main__":
    results = run_all_scenarios()
    
    print(f"\nCompleted testing {len(results)} scenarios")
    print("Check qed_test_results.json for detailed results")