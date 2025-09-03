#!/usr/bin/env python3
"""
QED Serviceability Calculator Excel File Analyzer
Identifies input and output cells for automated testing
"""

import openpyxl
from pathlib import Path

def analyze_qed_excel():
    """Analyze QED Excel file structure to identify key cells"""
    
    excel_path = Path(r"C:\Users\encou\Documents\Project MicroSass\Otium\qed_serviceability_calculator_3_28_may_2025_download_450.xlsm")
    
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        return None
    
    print(f"Analyzing QED Excel file: {excel_path.name}")
    
    try:
        # Load workbook with macros preserved
        wb = openpyxl.load_workbook(excel_path, keep_vba=True, data_only=False)
        print(f"SUCCESS: Loaded workbook with {len(wb.sheetnames)} sheets")
        
        # Print all worksheet names
        print(f"\nWorksheets:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
        
        # Analyze main worksheet (usually first one)
        main_ws = wb.active
        print(f"\nAnalyzing main worksheet: {main_ws.title}")
        
        # Find cells with specific keywords that likely indicate inputs/outputs
        input_keywords = ['income', 'salary', 'wage', 'hecs', 'help', 'dependents', 'rent', 'expenses', 'rate']
        output_keywords = ['borrowing', 'capacity', 'power', 'maximum', 'loan', 'amount', 'result']
        
        print(f"\nSearching for potential INPUT cells:")
        input_cells = {}
        
        for row in range(1, 100):  # Check first 100 rows
            for col in range(1, 26):  # Check first 26 columns (A-Z)
                cell = main_ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).lower()
                    for keyword in input_keywords:
                        if keyword in cell_text:
                            cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                            input_cells[keyword] = {
                                'cell': cell_ref,
                                'label': cell.value,
                                'adjacent_value': main_ws.cell(row=row, column=col+1).value
                            }
                            print(f"  INPUT {keyword.upper()}: {cell_ref} = '{cell.value}'")
                            break
        
        print(f"\nSearching for potential OUTPUT cells:")
        output_cells = {}
        
        for row in range(1, 100):
            for col in range(1, 26):
                cell = main_ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).lower()
                    for keyword in output_keywords:
                        if keyword in cell_text:
                            cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                            output_cells[keyword] = {
                                'cell': cell_ref,
                                'label': cell.value,
                                'adjacent_value': main_ws.cell(row=row, column=col+1).value
                            }
                            print(f"  OUTPUT {keyword.upper()}: {cell_ref} = '{cell.value}'")
                            break
        
        # Look for numeric values that might be results
        print(f"\nSearching for numeric values (potential results):")
        numeric_cells = []
        for row in range(1, 50):
            for col in range(1, 26):
                cell = main_ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and cell.value > 100000:  # Likely borrowing amounts
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    numeric_cells.append({
                        'cell': cell_ref,
                        'value': cell.value,
                        'formatted': f"${cell.value:,.0f}" if cell.value else "N/A"
                    })
        
        # Show largest numeric values (likely results)
        numeric_cells.sort(key=lambda x: x['value'] if x['value'] else 0, reverse=True)
        for cell_info in numeric_cells[:10]:
            print(f"  RESULT {cell_info['cell']}: {cell_info['formatted']}")
        
        # Save analysis results
        analysis_results = {
            'worksheets': wb.sheetnames,
            'input_cells': input_cells,
            'output_cells': output_cells,
            'potential_results': numeric_cells[:10]
        }
        
        wb.close()
        return analysis_results
        
    except Exception as e:
        print(f"ERROR: Error analyzing Excel file: {str(e)}")
        return None

def create_test_scenarios():
    """Create the 6 test scenarios for QED testing"""
    
    scenarios = [
        {
            'name': 'Scenario 1: Single, Low Income, Owner-Occupied',
            'primary_income': 65000,
            'secondary_income': 0,
            'dependents': 0,
            'hecs_primary': 0,
            'hecs_secondary': 0,
            'property_type': 'Owner-Occupied',
            'interest_rate': 5.5,
            'location': 'NSW 2000',
            'rental_income': 0,
            'current_rent': 650,
            'expected_range': (350000, 400000),
            'our_app_result': 312530
        },
        {
            'name': 'Scenario 2: Single, Medium Income, Investment',
            'primary_income': 95000,
            'secondary_income': 0,
            'dependents': 0,
            'hecs_primary': 25000,
            'hecs_secondary': 0,
            'property_type': 'Investment',
            'interest_rate': 5.8,
            'location': 'VIC 3000',
            'rental_income': 450 * 52,  # Weekly to annual
            'current_rent': 0,
            'expected_range': (600000, 700000),
            'our_app_result': 502553
        },
        {
            'name': 'Scenario 3: Couple, High Income, Owner-Occupied',
            'primary_income': 120000,
            'secondary_income': 85000,
            'dependents': 2,
            'hecs_primary': 35000,
            'hecs_secondary': 20000,
            'property_type': 'Owner-Occupied',
            'interest_rate': 5.5,
            'location': 'QLD 4000',
            'rental_income': 0,
            'current_rent': 650,
            'expected_range': (900000, 1000000),
            'our_app_result': 237413
        },
        {
            'name': 'Scenario 4: Couple, High Income, Investment',
            'primary_income': 140000,
            'secondary_income': 75000,
            'dependents': 0,
            'hecs_primary': 0,
            'hecs_secondary': 0,
            'property_type': 'Investment',
            'interest_rate': 5.8,
            'location': 'WA 6000',
            'rental_income': 650 * 52,  # Weekly to annual
            'current_rent': 400 * 52,   # Weekly to annual
            'expected_range': (1200000, 1500000),
            'our_app_result': 1060237
        },
        {
            'name': 'Scenario 5: Single, Very High Income, Investment',
            'primary_income': 180000,
            'secondary_income': 0,
            'dependents': 1,
            'hecs_primary': 45000,
            'hecs_secondary': 0,
            'property_type': 'Investment',
            'interest_rate': 5.8,
            'location': 'SA 5000',
            'rental_income': 800 * 52,  # Weekly to annual
            'current_rent': 0,
            'expected_range': (1500000, 2000000),
            'our_app_result': 660016
        },
        {
            'name': 'Scenario 6: Young Couple, Entry Level',
            'primary_income': 75000,
            'secondary_income': 60000,
            'dependents': 0,
            'hecs_primary': 15000,
            'hecs_secondary': 18000,
            'property_type': 'Owner-Occupied',
            'interest_rate': 5.5,
            'location': 'NSW 2650',
            'rental_income': 0,
            'current_rent': 650,
            'expected_range': (650000, 750000),
            'our_app_result': 425720
        }
    ]
    
    return scenarios

if __name__ == "__main__":
    print("QED Serviceability Calculator Analysis")
    print("=" * 50)
    
    # Analyze Excel structure
    results = analyze_qed_excel()
    
    if results:
        print(f"\nAnalysis complete!")
        print(f"Found {len(results['input_cells'])} potential input areas")
        print(f"Found {len(results['output_cells'])} potential output areas")
        print(f"Found {len(results['potential_results'])} potential result cells")
    else:
        print(f"\nAnalysis failed!")
    
    # Show test scenarios
    scenarios = create_test_scenarios()
    print(f"\nCreated {len(scenarios)} test scenarios ready for automation")
    
    print(f"\nNext Steps:")
    print(f"  1. Identify exact input cell locations from analysis above")
    print(f"  2. Identify exact output cell location")
    print(f"  3. Create automated test script")
    print(f"  4. Run all 6 scenarios automatically")